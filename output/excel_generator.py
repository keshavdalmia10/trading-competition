"""Excel report generator using openpyxl."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import COMPETITION_END, COMPETITION_START, SCORING_WEIGHTS
from data.models import (
    CatalystHunterOutput,
    FundamentalScreenerOutput,
    MacroAnalysis,
    PortfolioManagerOutput,
    RiskManagerOutput,
    ScoringRow,
    SentimentAnalystOutput,
    StockType,
    TechnicalAnalystOutput,
)
from orchestrator.message_bus import MessageBus
from tools.scoring import compute_composite_score


# ── Styles ────────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
GOOD_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _apply_header_style(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_width: int = 12, max_width: int = 50) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    """Write a table with headers and return the next empty row."""
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=header)
    _apply_header_style(ws, start_row, len(headers))

    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

    return start_row + len(rows) + 2


# ── Sheet Builders ────────────────────────────────────────────────────────

def _build_portfolio_summary(wb: Workbook, portfolio: PortfolioManagerOutput) -> None:
    ws = wb.active
    ws.title = "Portfolio Summary"

    # Title
    ws.cell(row=1, column=1, value="Trading Competition Portfolio").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Period: {COMPETITION_START} to {COMPETITION_END}")
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws.cell(row=4, column=1, value=f"Evolution: {portfolio.evolution_count} | Revolution: {portfolio.revolution_count}")

    headers = [
        "Ticker", "Name", "Type", "Sector", "Composite Score",
        "Weight %", "Entry Strategy", "Exit Strategy", "Thesis",
    ]
    rows = []
    for s in sorted(portfolio.stocks, key=lambda x: x.composite_score, reverse=True):
        rows.append([
            s.ticker, s.name, s.stock_type.value.title(), s.sector,
            round(s.composite_score, 1), round(s.weight_pct, 1),
            s.entry_strategy, s.exit_strategy, s.thesis,
        ])

    next_row = _write_table(ws, headers, rows, start_row=6)

    # Portfolio metadata
    ws.cell(row=next_row, column=1, value="Portfolio Rationale").font = Font(bold=True, size=11)
    ws.cell(row=next_row + 1, column=1, value=portfolio.portfolio_rationale).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=9)

    next_row += 3
    ws.cell(row=next_row, column=1, value="Key Risks").font = Font(bold=True)
    for i, risk in enumerate(portfolio.key_risks):
        ws.cell(row=next_row + 1 + i, column=1, value=f"• {risk}")

    next_row += len(portfolio.key_risks) + 2
    ws.cell(row=next_row, column=1, value="Key Catalysts").font = Font(bold=True)
    for i, cat in enumerate(portfolio.key_catalysts):
        ws.cell(row=next_row + 1 + i, column=1, value=f"• {cat}")

    _auto_width(ws)
    # Make thesis column wider
    ws.column_dimensions["I"].width = 60


def _build_macro_overview(wb: Workbook, macro: MacroAnalysis) -> None:
    ws = wb.create_sheet("Macro Overview")

    ws.cell(row=1, column=1, value="Macroeconomic Overview").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Market Regime: {macro.regime.value.replace('_', ' ').title()}")
    ws.cell(row=2, column=1).font = Font(bold=True, size=12)
    ws.cell(row=3, column=1, value=f"Macro Score: {macro.macro_score}/100")

    ws.cell(row=5, column=1, value="Regime Rationale").font = Font(bold=True)
    ws.cell(row=6, column=1, value=macro.regime_rationale).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=4)

    # Favored / Avoided sectors
    ws.cell(row=8, column=1, value="Favored Sectors").font = Font(bold=True)
    ws.cell(row=8, column=3, value="Avoided Sectors").font = Font(bold=True)
    for i, s in enumerate(macro.favored_sectors):
        ws.cell(row=9 + i, column=1, value=s).fill = GOOD_FILL
    for i, s in enumerate(macro.avoided_sectors):
        ws.cell(row=9 + i, column=3, value=s).fill = BAD_FILL

    # Indicators table
    start = 9 + max(len(macro.favored_sectors), len(macro.avoided_sectors)) + 2
    headers = ["Indicator", "Value", "Interpretation"]
    rows = [[ind.name, ind.value, ind.interpretation] for ind in macro.indicators]
    next_row = _write_table(ws, headers, rows, start_row=start)

    # Key events
    ws.cell(row=next_row, column=1, value="Key Upcoming Events").font = Font(bold=True)
    for i, event in enumerate(macro.key_events):
        ws.cell(row=next_row + 1 + i, column=1, value=f"• {event}")

    next_row += len(macro.key_events) + 2
    ws.cell(row=next_row, column=1, value="Summary").font = Font(bold=True)
    ws.cell(row=next_row + 1, column=1, value=macro.summary).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=4)

    _auto_width(ws)


def _build_stock_sheet(
    wb: Workbook,
    ticker: str,
    portfolio: PortfolioManagerOutput,
    fundamental: FundamentalScreenerOutput | None,
    technical: TechnicalAnalystOutput | None,
    catalyst: CatalystHunterOutput | None,
    sentiment: SentimentAnalystOutput | None,
    risk: RiskManagerOutput | None,
) -> None:
    """Build a detailed sheet for a single stock."""
    ws = wb.create_sheet(ticker)
    stock = next((s for s in portfolio.stocks if s.ticker == ticker), None)
    if not stock:
        return

    # Header
    ws.cell(row=1, column=1, value=f"{ticker} — {stock.name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Type: {stock.stock_type.value.title()} | Sector: {stock.sector} | Weight: {stock.weight_pct}%")
    ws.cell(row=3, column=1, value=f"Composite Score: {stock.composite_score}/100").font = Font(bold=True, size=12)

    # Score breakdown
    row = 5
    ws.cell(row=row, column=1, value="Score Breakdown").font = Font(bold=True, size=11)
    scores = [
        ["Fundamental", stock.fundamental_score, f"{SCORING_WEIGHTS['fundamental']*100}%"],
        ["Technical", stock.technical_score, f"{SCORING_WEIGHTS['technical']*100}%"],
        ["Catalyst", stock.catalyst_score, f"{SCORING_WEIGHTS['catalyst']*100}%"],
        ["Sentiment", stock.sentiment_score, f"{SCORING_WEIGHTS['sentiment']*100}%"],
        ["Risk", stock.risk_score, f"{SCORING_WEIGHTS['risk']*100}%"],
        ["COMPOSITE", stock.composite_score, "100%"],
    ]
    row = _write_table(ws, ["Category", "Score (0-100)", "Weight"], scores, start_row=row + 1)

    # Investment Thesis
    ws.cell(row=row, column=1, value="Investment Thesis").font = Font(bold=True, size=11)
    ws.cell(row=row + 1, column=1, value=stock.thesis).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=4)
    row += 3

    # Entry / Exit
    ws.cell(row=row, column=1, value="Entry Strategy").font = Font(bold=True)
    ws.cell(row=row, column=2, value=stock.entry_strategy)
    ws.cell(row=row + 1, column=1, value="Exit Strategy").font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value=stock.exit_strategy)
    row += 3

    # Fundamentals section
    fund_data = None
    if fundamental:
        fund_data = next((c for c in fundamental.candidates if c.ticker == ticker), None)
    if fund_data:
        ws.cell(row=row, column=1, value="Fundamentals").font = Font(bold=True, size=11)
        row += 1
        for label, val in [
            ("P/E Ratio", fund_data.pe_ratio),
            ("PEG Ratio", fund_data.peg_ratio),
            ("Revenue Growth YoY", f"{fund_data.revenue_growth_yoy:.1%}" if fund_data.revenue_growth_yoy else "N/A"),
            ("EPS Growth YoY", f"{fund_data.eps_growth_yoy:.1%}" if fund_data.eps_growth_yoy else "N/A"),
            ("ROE", f"{fund_data.roe:.1%}" if fund_data.roe else "N/A"),
            ("Earnings Surprise", f"{fund_data.earnings_surprise_pct:.1f}%" if fund_data.earnings_surprise_pct else "N/A"),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1
        ws.cell(row=row, column=1, value="Rationale").font = Font(bold=True)
        ws.cell(row=row, column=2, value=fund_data.rationale).alignment = WRAP_ALIGNMENT
        row += 2

    # Technical section
    tech_data = None
    if technical:
        tech_data = next((t for t in technical.analyses if t.ticker == ticker), None)
    if tech_data:
        ws.cell(row=row, column=1, value="Technical Analysis").font = Font(bold=True, size=11)
        row += 1
        for label, val in [
            ("Current Price", tech_data.current_price),
            ("RSI (14)", tech_data.rsi_14),
            ("MACD Signal", tech_data.macd_signal),
            ("Above SMA 20", tech_data.above_sma_20),
            ("Above SMA 50", tech_data.above_sma_50),
            ("Volume Trend", tech_data.volume_trend),
            ("Support", tech_data.support_level),
            ("Resistance", tech_data.resistance_level),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1
        ws.cell(row=row, column=1, value="Rationale").font = Font(bold=True)
        ws.cell(row=row, column=2, value=tech_data.rationale).alignment = WRAP_ALIGNMENT
        row += 2

    # Catalyst section
    cat_data = None
    if catalyst:
        cat_data = next((c for c in catalyst.analyses if c.ticker == ticker), None)
    if cat_data and cat_data.catalysts:
        ws.cell(row=row, column=1, value="Catalysts").font = Font(bold=True, size=11)
        row += 1
        cat_headers = ["Event", "Date", "Impact", "Direction", "Description"]
        cat_rows = [
            [c.event, c.date or "TBD", c.impact.value, c.direction.value, c.description]
            for c in cat_data.catalysts
        ]
        row = _write_table(ws, cat_headers, cat_rows, start_row=row)

    # Sentiment section
    sent_data = None
    if sentiment:
        sent_data = next((s for s in sentiment.analyses if s.ticker == ticker), None)
    if sent_data:
        ws.cell(row=row, column=1, value="Sentiment").font = Font(bold=True, size=11)
        row += 1
        for label, val in [
            ("Overall Sentiment", sent_data.overall_sentiment),
            ("Analyst Consensus", sent_data.analyst_consensus),
            ("Target Price", sent_data.analyst_target_price),
            ("Recent Upgrades", sent_data.recent_upgrades),
            ("Recent Downgrades", sent_data.recent_downgrades),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1
        if sent_data.key_headlines:
            ws.cell(row=row, column=1, value="Key Headlines").font = Font(bold=True)
            row += 1
            for h in sent_data.key_headlines[:5]:
                ws.cell(row=row, column=1, value=f"• {h}").alignment = WRAP_ALIGNMENT
                row += 1
        row += 1

    # Risk section
    risk_data = None
    if risk:
        risk_data = next((r for r in risk.analyses if r.ticker == ticker), None)
    if risk_data:
        ws.cell(row=row, column=1, value="Risk Profile").font = Font(bold=True, size=11)
        row += 1
        for label, val in [
            ("Beta", risk_data.beta),
            ("Annualized Volatility", f"{risk_data.volatility_annualized:.1%}" if risk_data.volatility_annualized else "N/A"),
            ("Max Drawdown (90d)", f"{risk_data.max_drawdown_90d:.1%}" if risk_data.max_drawdown_90d else "N/A"),
            ("VaR (95%)", f"{risk_data.value_at_risk_95:.2%}" if risk_data.value_at_risk_95 else "N/A"),
            ("Sharpe Ratio", risk_data.sharpe_ratio),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            row += 1

    _auto_width(ws)


def _build_scoring_matrix(
    wb: Workbook,
    fundamental: FundamentalScreenerOutput | None,
    technical: TechnicalAnalystOutput | None,
    catalyst: CatalystHunterOutput | None,
    sentiment: SentimentAnalystOutput | None,
    risk: RiskManagerOutput | None,
    portfolio: PortfolioManagerOutput,
) -> None:
    ws = wb.create_sheet("Scoring Matrix")
    ws.cell(row=1, column=1, value="Candidate Scoring Matrix").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Weights: Fund={SCORING_WEIGHTS['fundamental']:.0%} Tech={SCORING_WEIGHTS['technical']:.0%} Cat={SCORING_WEIGHTS['catalyst']:.0%} Sent={SCORING_WEIGHTS['sentiment']:.0%} Risk={SCORING_WEIGHTS['risk']:.0%}")

    selected_tickers = {s.ticker for s in portfolio.stocks}

    # Build scoring rows
    rows_data = []
    if fundamental:
        for c in fundamental.candidates:
            tech_score = 50.0
            cat_score = 50.0
            sent_score = 50.0
            risk_score = 50.0

            if technical:
                t = next((x for x in technical.analyses if x.ticker == c.ticker), None)
                if t:
                    tech_score = t.technical_score
            if catalyst:
                ca = next((x for x in catalyst.analyses if x.ticker == c.ticker), None)
                if ca:
                    cat_score = ca.catalyst_score
            if sentiment:
                s = next((x for x in sentiment.analyses if x.ticker == c.ticker), None)
                if s:
                    sent_score = s.sentiment_score
            if risk:
                r = next((x for x in risk.analyses if x.ticker == c.ticker), None)
                if r:
                    risk_score = r.risk_score

            # Gather algo scores
            momentum_algo = None
            risk_algo = None
            if technical:
                t = next((x for x in technical.analyses if x.ticker == c.ticker), None)
                if t:
                    momentum_algo = t.momentum_score_algo
            if risk:
                r2 = next((x for x in risk.analyses if x.ticker == c.ticker), None)
                if r2:
                    risk_algo = r2.risk_adjusted_score_algo

            composite = compute_composite_score(
                c.fundamental_score, tech_score, cat_score, sent_score, risk_score
            )
            rows_data.append(ScoringRow(
                ticker=c.ticker,
                name=c.name,
                sector=c.sector,
                stock_type=c.stock_type,
                fundamental_score=c.fundamental_score,
                technical_score=tech_score,
                catalyst_score=cat_score,
                sentiment_score=sent_score,
                risk_score=risk_score,
                composite_score=composite,
                momentum_score_algo=momentum_algo,
                quality_score_algo=c.quality_score_algo,
                earnings_surprise_score_algo=c.earnings_surprise_score_algo,
                risk_adjusted_score_algo=risk_algo,
                selected=c.ticker in selected_tickers,
            ))

    rows_data.sort(key=lambda r: r.composite_score, reverse=True)

    headers = [
        "Ticker", "Name", "Sector", "Type",
        "Fund", "Fund Algo (Qual)", "Fund Algo (Earn)",
        "Tech", "Tech Algo (Mom)",
        "Catalyst", "Sentiment",
        "Risk", "Risk Algo",
        "Composite", "Selected",
    ]
    rows = []
    for r in rows_data:
        rows.append([
            r.ticker, r.name, r.sector, r.stock_type.value.title(),
            round(r.fundamental_score, 1),
            round(r.quality_score_algo, 1) if r.quality_score_algo is not None else "N/A",
            round(r.earnings_surprise_score_algo, 1) if r.earnings_surprise_score_algo is not None else "N/A",
            round(r.technical_score, 1),
            round(r.momentum_score_algo, 1) if r.momentum_score_algo is not None else "N/A",
            round(r.catalyst_score, 1), round(r.sentiment_score, 1),
            round(r.risk_score, 1),
            round(r.risk_adjusted_score_algo, 1) if r.risk_adjusted_score_algo is not None else "N/A",
            round(r.composite_score, 1),
            "YES" if r.selected else "",
        ])

    next_row = _write_table(ws, headers, rows, start_row=4)

    # Color the selected rows
    for row_idx in range(5, 5 + len(rows)):
        if ws.cell(row=row_idx, column=15).value == "YES":
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).fill = GOOD_FILL

    _auto_width(ws)


def _build_risk_analysis(wb: Workbook, risk: RiskManagerOutput | None) -> None:
    ws = wb.create_sheet("Risk Analysis")
    ws.cell(row=1, column=1, value="Risk Analysis").font = Font(bold=True, size=14)

    if not risk:
        ws.cell(row=3, column=1, value="No risk data available")
        return

    ws.cell(row=2, column=1, value=f"Portfolio Beta: {risk.portfolio_beta or 'N/A'}")

    # Risk metrics table
    headers = ["Ticker", "Beta", "Volatility (Ann.)", "Max Drawdown", "VaR 95%", "Sharpe", "Risk Score", "Weight"]
    rows = []
    for r in sorted(risk.analyses, key=lambda x: x.risk_score, reverse=True):
        rows.append([
            r.ticker,
            round(r.beta, 2) if r.beta else "N/A",
            f"{r.volatility_annualized:.1%}" if r.volatility_annualized else "N/A",
            f"{r.max_drawdown_90d:.1%}" if r.max_drawdown_90d else "N/A",
            f"{r.value_at_risk_95:.2%}" if r.value_at_risk_95 else "N/A",
            round(r.sharpe_ratio, 2) if r.sharpe_ratio else "N/A",
            round(r.risk_score, 1),
            f"{r.suggested_weight:.1%}",
        ])

    next_row = _write_table(ws, headers, rows, start_row=4)

    # High correlations
    if risk.high_correlations:
        ws.cell(row=next_row, column=1, value="High Correlation Pairs (>0.7)").font = Font(bold=True, size=11)
        corr_headers = ["Stock A", "Stock B", "Correlation"]
        corr_rows = [[c.ticker_a, c.ticker_b, round(c.correlation, 3)] for c in risk.high_correlations]
        next_row = _write_table(ws, corr_headers, corr_rows, start_row=next_row + 1)

    ws.cell(row=next_row, column=1, value="Diversification Notes").font = Font(bold=True)
    ws.cell(row=next_row + 1, column=1, value=risk.diversification_notes).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=next_row + 1, start_column=1, end_row=next_row + 1, end_column=8)

    _auto_width(ws)


def _build_agent_summaries(
    wb: Workbook,
    macro: MacroAnalysis | None,
    fundamental: FundamentalScreenerOutput | None,
    technical: TechnicalAnalystOutput | None,
    catalyst: CatalystHunterOutput | None,
    sentiment: SentimentAnalystOutput | None,
    risk: RiskManagerOutput | None,
    portfolio: PortfolioManagerOutput,
) -> None:
    """Build a sheet showing each agent's overall analysis summary."""
    ws = wb.create_sheet("Agent Summaries")
    ws.cell(row=1, column=1, value="Agent-by-Agent Analysis Report").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Each agent's summary, key findings, and reasoning").font = Font(italic=True, color="666666")
    row = 4

    # Helper to write an agent section
    def _section(title: str, summary: str, details: list[tuple[str, str]] | None = None) -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13, color="1F4E79")
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = SUBHEADER_FILL
        row += 1
        ws.cell(row=row, column=1, value="Summary:").font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value=summary).alignment = WRAP_ALIGNMENT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2
        if details:
            for label, value in details:
                ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row, column=2, value=value).alignment = WRAP_ALIGNMENT
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
                row += 1
        row += 1
        return row

    # 1. Macro Analyst
    if macro:
        details = [
            ("Market Regime", f"{macro.regime.value.replace('_', ' ').title()} (Score: {macro.macro_score}/100)"),
            ("Regime Rationale", macro.regime_rationale),
            ("Favored Sectors", ", ".join(macro.favored_sectors)),
            ("Avoided Sectors", ", ".join(macro.avoided_sectors)),
            ("Key Events", "\n".join(f"- {e}" for e in macro.key_events)),
        ]
        _section("Phase 0A: Macro Analyst", macro.summary, details)

    # 2. Fundamental Screener
    if fundamental:
        top5 = sorted(fundamental.candidates, key=lambda c: c.fundamental_score, reverse=True)[:5]
        top5_str = "\n".join(f"- {c.ticker} ({c.name}): {c.fundamental_score}/100 — {c.rationale}" for c in top5)
        details = [
            ("Candidates Screened", str(len(fundamental.candidates))),
            ("Top 5 by Fund. Score", top5_str),
        ]
        _section("Phase 0B: Fundamental Screener", fundamental.screening_summary, details)

    # 3. Technical Analyst
    if technical:
        top5 = sorted(technical.analyses, key=lambda t: t.technical_score, reverse=True)[:5]
        top5_str = "\n".join(f"- {t.ticker}: {t.technical_score}/100 (RSI={t.rsi_14}, MACD={t.macd_signal}) — {t.rationale}" for t in top5)
        details = [
            ("Tickers Analyzed", str(len(technical.analyses))),
            ("Top 5 by Tech. Score", top5_str),
        ]
        _section("Phase 1A: Technical Analyst", technical.summary, details)

    # 4. Catalyst Hunter
    if catalyst:
        top5 = sorted(catalyst.analyses, key=lambda c: c.catalyst_score, reverse=True)[:5]
        top5_lines = []
        for c in top5:
            cat_names = ", ".join(cat.event for cat in c.catalysts[:3]) if c.catalysts else "None"
            top5_lines.append(f"- {c.ticker}: {c.catalyst_score}/100 (Catalysts: {cat_names}) — {c.rationale}")
        details = [
            ("Tickers Analyzed", str(len(catalyst.analyses))),
            ("Top 5 by Catalyst Score", "\n".join(top5_lines)),
        ]
        _section("Phase 1B: Catalyst Hunter", catalyst.summary, details)

    # 5. Sentiment Analyst
    if sentiment:
        top5 = sorted(sentiment.analyses, key=lambda s: s.sentiment_score, reverse=True)[:5]
        top5_str = "\n".join(
            f"- {s.ticker}: {s.sentiment_score}/100 (Sentiment={s.overall_sentiment:+.2f}, Consensus={s.analyst_consensus}) — {s.rationale}"
            for s in top5
        )
        details = [
            ("Tickers Analyzed", str(len(sentiment.analyses))),
            ("Top 5 by Sentiment Score", top5_str),
        ]
        _section("Phase 1C: Sentiment Analyst", sentiment.summary, details)

    # 6. Risk Manager
    if risk:
        top5 = sorted(risk.analyses, key=lambda r: r.risk_score, reverse=True)[:5]
        top5_lines = []
        for r in top5:
            vol_str = f"{r.volatility_annualized:.1%}" if r.volatility_annualized else "N/A"
            regime = r.vol_regime or "N/A"
            top5_lines.append(f"- {r.ticker}: {r.risk_score}/100 (Vol={vol_str}, Regime={regime}, Beta={r.beta}) — {r.rationale}")
        corr_str = ", ".join(f"{c.ticker_a}/{c.ticker_b}={c.correlation:.2f}" for c in risk.high_correlations[:5]) if risk.high_correlations else "None above 0.7"
        details = [
            ("Tickers Analyzed", str(len(risk.analyses))),
            ("Portfolio Beta", str(risk.portfolio_beta or "N/A")),
            ("High Correlations", corr_str),
            ("Diversification Notes", risk.diversification_notes),
            ("Top 5 by Risk Score", "\n".join(top5_lines)),
        ]
        _section("Phase 2A: Risk Manager", risk.summary, details)

    # 7. Portfolio Manager
    selected = sorted(portfolio.stocks, key=lambda s: s.composite_score, reverse=True)
    stock_lines = []
    for s in selected:
        stock_lines.append(f"- {s.ticker} ({s.name}): {s.weight_pct}% weight, Score={s.composite_score}/100 — {s.thesis}")
    details = [
        ("Stocks Selected", str(len(portfolio.stocks))),
        ("Evolution / Revolution", f"{portfolio.evolution_count} / {portfolio.revolution_count}"),
        ("Sector Breakdown", ", ".join(f"{k}: {v}" for k, v in portfolio.sector_breakdown.items())),
        ("Key Risks", "\n".join(f"- {r}" for r in portfolio.key_risks)),
        ("Key Catalysts", "\n".join(f"- {c}" for c in portfolio.key_catalysts)),
        ("Portfolio Selections", "\n".join(stock_lines)),
    ]
    _section("Phase 2B: Portfolio Manager", portfolio.portfolio_rationale, details)

    # Set column widths
    ws.column_dimensions["A"].width = 25
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 30


def _build_rationale_matrix(
    wb: Workbook,
    portfolio: PortfolioManagerOutput,
    fundamental: FundamentalScreenerOutput | None,
    technical: TechnicalAnalystOutput | None,
    catalyst: CatalystHunterOutput | None,
    sentiment: SentimentAnalystOutput | None,
    risk: RiskManagerOutput | None,
) -> None:
    """Build a sheet with per-stock rationales from every agent side by side."""
    ws = wb.create_sheet("Rationale Matrix")
    ws.cell(row=1, column=1, value="Per-Stock Rationale Comparison").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value="Why each agent scored each stock the way it did").font = Font(italic=True, color="666666")

    headers = [
        "Ticker", "Score",
        "Fundamental Rationale", "Fund Score",
        "Technical Rationale", "Tech Score",
        "Catalyst Rationale", "Cat Score",
        "Sentiment Rationale", "Sent Score",
        "Risk Rationale", "Risk Score",
        "Portfolio Thesis",
    ]
    _apply_header_style(ws, 4, len(headers))
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=4, column=col_idx, value=header)

    # Build rows for each selected stock
    selected = sorted(portfolio.stocks, key=lambda s: s.composite_score, reverse=True)
    for row_idx, stock in enumerate(selected, 5):
        ticker = stock.ticker

        # Gather rationales from each agent
        fund_rat = fund_score = ""
        if fundamental:
            fd = next((c for c in fundamental.candidates if c.ticker == ticker), None)
            if fd:
                fund_rat = fd.rationale
                fund_score = round(fd.fundamental_score, 1)

        tech_rat = tech_score = ""
        if technical:
            td = next((t for t in technical.analyses if t.ticker == ticker), None)
            if td:
                tech_rat = td.rationale
                tech_score = round(td.technical_score, 1)

        cat_rat = cat_score = ""
        if catalyst:
            cd = next((c for c in catalyst.analyses if c.ticker == ticker), None)
            if cd:
                cat_rat = cd.rationale
                cat_score = round(cd.catalyst_score, 1)

        sent_rat = sent_score = ""
        if sentiment:
            sd = next((s for s in sentiment.analyses if s.ticker == ticker), None)
            if sd:
                sent_rat = sd.rationale
                sent_score = round(sd.sentiment_score, 1)

        risk_rat = risk_score = ""
        if risk:
            rd = next((r for r in risk.analyses if r.ticker == ticker), None)
            if rd:
                risk_rat = rd.rationale
                risk_score = round(rd.risk_score, 1)

        values = [
            ticker, round(stock.composite_score, 1),
            fund_rat, fund_score,
            tech_rat, tech_score,
            cat_rat, cat_score,
            sent_rat, sent_score,
            risk_rat, risk_score,
            stock.thesis,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

        # Color score cells
        for score_col in [2, 4, 6, 8, 10, 12]:
            cell = ws.cell(row=row_idx, column=score_col)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 70:
                    cell.fill = GOOD_FILL
                elif cell.value >= 50:
                    cell.fill = WARN_FILL
                else:
                    cell.fill = BAD_FILL

    # Set column widths — rationale columns wider
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    for col_letter in ["C", "E", "G", "I", "K", "M"]:
        ws.column_dimensions[col_letter].width = 45
    for col_letter in ["D", "F", "H", "J", "L"]:
        ws.column_dimensions[col_letter].width = 10


# ── Main Generator ────────────────────────────────────────────────────────

def generate_report(bus: MessageBus) -> Path:
    """Generate the full multi-sheet Excel report from message bus data."""
    wb = Workbook()

    portfolio: PortfolioManagerOutput | None = bus.get("portfolio_manager")
    macro: MacroAnalysis | None = bus.get("macro_analyst")
    fundamental: FundamentalScreenerOutput | None = bus.get("fundamental_screener")
    technical: TechnicalAnalystOutput | None = bus.get("technical_analyst")
    catalyst: CatalystHunterOutput | None = bus.get("catalyst_hunter")
    sentiment: SentimentAnalystOutput | None = bus.get("sentiment_analyst")
    risk: RiskManagerOutput | None = bus.get("risk_manager")

    if not portfolio:
        logger.error("No portfolio data on bus — cannot generate report!")
        raise ValueError("Portfolio manager output not found on message bus")

    # Sheet 1: Portfolio Summary
    _build_portfolio_summary(wb, portfolio)

    # Sheet 2: Macro Overview
    if macro:
        _build_macro_overview(wb, macro)

    # Sheets 3-12: Individual stock sheets
    for stock in portfolio.stocks:
        _build_stock_sheet(wb, stock.ticker, portfolio, fundamental, technical, catalyst, sentiment, risk)

    # Sheet: Scoring Matrix
    _build_scoring_matrix(wb, fundamental, technical, catalyst, sentiment, risk, portfolio)

    # Sheet: Risk Analysis
    _build_risk_analysis(wb, risk)

    # Sheet: Agent Summaries (step-by-step reasoning)
    _build_agent_summaries(wb, macro, fundamental, technical, catalyst, sentiment, risk, portfolio)

    # Sheet: Rationale Matrix (per-stock rationale comparison)
    _build_rationale_matrix(wb, portfolio, fundamental, technical, catalyst, sentiment, risk)

    # Save
    output_dir = Path(__file__).parent / "reports"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"trading_competition_{timestamp}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    logger.info(f"Excel report saved to {filepath}")

    return filepath
